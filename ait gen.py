import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from openpyxl import load_workbook, Workbook
import datetime
from tkcalendar import DateEntry

# --- Allergen Data ---
ALLERGENS = [
    # Mold Group
    {"name": "Aspergillus", "group": "Mold", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Tree", "Grass", "Weed", "Other"]},
    {"name": "Alternaria", "group": "Mold", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Tree", "Grass", "Weed", "Other"]},
    {"name": "Cladosporium", "group": "Mold", "min_volume": 1.0, "max_volume": 1.0, "incompatible_groups": ["Tree", "Grass", "Weed", "Other"]},
    {"name": "Penicillium", "group": "Mold", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Tree", "Grass", "Weed", "Other"]},

    # Tree Group
    {"name": "Ash", "group": "Tree", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Mold", "Other"]},
    {"name": "Birch (Oak)", "group": "Tree", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Mold", "Other"]},
    {"name": "Cedar", "group": "Tree", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Mold", "Other"]},
    {"name": "Hackberry (Elm)", "group": "Tree", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Mold", "Other"]},
    {"name": "Maple", "group": "Tree", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Mold", "Other"]},
    {"name": "Sycamore", "group": "Tree", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Mold", "Other"]},
    {"name": "Walnut (Pecan)", "group": "Tree", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Mold", "Other"]},
    {"name": "Willow (Cottonwood)", "group": "Tree", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Mold", "Other"]},
    {"name": "Mulberry", "group": "Tree", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Mold", "Other"]},

    # Grass Group
    {"name": "Timothy", "group": "Grass", "min_volume": 0.1, "max_volume": 0.4, "incompatible_groups": ["Mold", "Other"]},
    {"name": "Johnson", "group": "Grass", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Mold", "Other"]},
    {"name": "Bermuda", "group": "Grass", "min_volume": 0.3, "max_volume": 1.5, "incompatible_groups": ["Mold", "Other"]},

    # Weed Group
    {"name": "Cocklebur", "group": "Weed", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Mold", "Other"]},
    {"name": "Yellow Dock (Sheep Sorrel)", "group": "Weed", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Mold", "Other"]},
    {"name": "Kochia (Firebush)", "group": "Weed", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Mold", "Other"]},
    {"name": "Lamb's Quarter", "group": "Weed", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Mold", "Other"]},
    {"name": "Mugwort", "group": "Weed", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Mold", "Other"]},
    {"name": "Pigweed", "group": "Weed", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Mold", "Other"]},
    {"name": "English Plantain", "group": "Weed", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Mold", "Other"]},
    {"name": "Russian Thistle", "group": "Weed", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Mold", "Other"]},
    {"name": "Ragweed", "group": "Weed", "min_volume": 0.3, "max_volume": 0.6, "incompatible_groups": ["Mold", "Other"]},

    # Other Group
    {"name": "Cat", "group": "Other", "min_volume": 1.0, "max_volume": 4.0, "incompatible_groups": ["Tree", "Grass", "Weed", "Mold"]},
    {"name": "Dog - UF", "group": "Other", "min_volume": 1.0, "max_volume": 1.0, "incompatible_groups": ["Tree", "Grass", "Weed", "Mold"]},
    {"name": "Dog - Epithelium", "group": "Other", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Tree", "Grass", "Weed", "Mold"]},
    {"name": "Mouse", "group": "Other", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Tree", "Grass", "Weed", "Mold"]},
    {"name": "Horse", "group": "Other", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Tree", "Grass", "Weed", "Mold"]},
    {"name": "Amer. Cockroach", "group": "Other", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Tree", "Grass", "Weed"]},
    {"name": "Ger. Cockroach", "group": "Other", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Tree", "Grass", "Weed"]},
    {"name": "Dust Mite Mix", "group": "Other", "min_volume": 0.5, "max_volume": 2.0, "incompatible_groups": ["Tree", "Grass", "Weed"]},

     # Venom Group
    {"name": "Honey Bee", "group": "Venom", "min_volume": 1.0, "max_volume": 1.0, "incompatible_groups": []},
    {"name": "Yellow Jacket", "group": "Venom", "min_volume": 1.0, "max_volume": 1.0, "incompatible_groups": []},
    {"name": "Yellow Faced Hornet", "group": "Venom", "min_volume": 1.0, "max_volume": 1.0, "incompatible_groups": []},
    {"name": "White Faced Hornet", "group": "Venom", "min_volume": 1.0, "max_volume": 1.0, "incompatible_groups": []},
    {"name": "Wasp", "group": "Venom", "min_volume": 1.0, "max_volume": 1.0, "incompatible_groups": []},
]

class Vial:
    """Represents a single allergy vial."""

    def __init__(self, label):
        self.label = label
        self.allergens = {}  # {allergen_name: volume}
        self.current_volume = 0.0

    def add_allergen(self, allergen_name, volume):
        """Adds an allergen to the vial if compatible and within volume limits.

        Args:
            allergen_name: The name of the allergen.
            volume: The volume of the allergen to add.

        Returns:
            True if the allergen was added successfully, False otherwise.
        """

        allergen_data = next((a for a in ALLERGENS if a["name"] == allergen_name), None)
        if not allergen_data:
            return False  # Allergen not found

        if not self.is_compatible(allergen_data):
            return False  # Incompatible allergen

        if self.current_volume + volume > 5.0:
            return False  # Exceeds volume limit

        if not (allergen_data["min_volume"] <= volume <= allergen_data["max_volume"]):
            return False  # Volume out of range.

        self.allergens[allergen_name] = volume
        self.current_volume += volume
        return True

    def remaining_volume(self):
        """Calculates the remaining volume in the vial."""
        return 5.0 - self.current_volume

    def is_compatible(self, allergen_data):
        """Checks if an allergen is compatible with the current vial contents."""
        current_groups = {a["group"] for a in ALLERGENS if a["name"] in self.allergens}
        for group in current_groups:
            if group in allergen_data["incompatible_groups"]:
                return False
        if allergen_data["group"] in {a["group"] for a in ALLERGENS for current_a in self.allergens if a["name"] == current_a}:
          return False
        return True

    def get_contents_string(self):
        """Returns a formatted string of the vial's contents."""
        contents = []
        for allergen, volume in self.allergens.items():
            contents.append(f"  - {allergen}: {volume:.2f} mL")
        contents.append(f"  - Diluent: {self.remaining_volume():.2f} mL")
        return "\n".join(contents)


def save_patient_data(patient_data, selected_allergens):
    """Saves patient data and selected allergens to the Excel file."""
    try:
        try:
            workbook = load_workbook(filename="patient_data.xlsx")
            sheet = workbook["Sheet1"]
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Sheet1"
            sheet.append(["Patient Name", "Date of Birth", "MRN", "Street Address", "City", "State", "Phone Number", "Allergens"])

        # Check for existing patient (using MRN)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[2] == patient_data['mrn']:
                response = messagebox.askyesno("Patient Exists",
                                             f"Patient with MRN '{patient_data['mrn']}' already exists. Overwrite?")
                if response:
                    row_num_to_delete = 0
                    for i,r in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                        if r[2] == patient_data['mrn']:
                            row_num_to_delete = i
                            break
                    if row_num_to_delete:
                        sheet.delete_rows(row_num_to_delete)
                else:
                    return

        # Append/Update new patient data, including allergens
        sheet.append([
            patient_data['patient_name'],
            patient_data['dob'].strftime("%m-%d-%Y"),
            patient_data['mrn'],
            patient_data['address'],
            patient_data['city'],
            patient_data['state'],
            patient_data['phone'],
            selected_allergens  # Comma-separated string of allergens
        ])
        workbook.save(filename="patient_data.xlsx")
        messagebox.showinfo("Success", f"Patient '{patient_data['patient_name']}' saved successfully.")

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while saving: {e}")

def load_patient():
    """Opens a new window to load an existing patient."""

    def load_selected_patient():
        selected_patient_str = patient_select_var.get()
        if not selected_patient_str:
            return

        try:
            # Split the string into name and DOB parts
            name_part, dob_part = selected_patient_str.rsplit(" ", 1)  # Split on the last space
            dob_part = dob_part.strip() #Remove extra space if there is any.
            name_part = name_part.strip()

            # Convert the DOB string to a date object for comparison
            dob_to_match = datetime.datetime.strptime(dob_part, "%m-%d-%Y").date()


            workbook = load_workbook(filename="patient_data.xlsx", read_only=True)
            sheet = workbook["Sheet1"]

            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Compare both name and DOB
                try:
                    row_dob = datetime.datetime.strptime(row[1], "%m-%d-%Y").date()
                except ValueError:
                    #Handle possible date conversion issues (e.g., bad data in Excel)
                    continue  # Skip this row if the date is invalid

                if row[0] == name_part and row_dob == dob_to_match:
                    # Populate the main window's fields
                    patient_name_entry.delete(0, tk.END)
                    patient_name_entry.insert(0, row[0])

                    # Date of Birth
                    dob_date = datetime.datetime.strptime(row[1], "%m-%d-%Y").date()
                    dob_entry.set_date(dob_date)

                    mrn_entry.delete(0, tk.END)
                    mrn_entry.insert(0, row[2])
                    address_entry.delete(0, tk.END)
                    address_entry.insert(0, row[3])
                    city_entry.delete(0, tk.END)
                    city_entry.insert(0, row[4])
                    state_entry.delete(0, tk.END)
                    state_entry.insert(0, row[5])
                    phone_entry.delete(0, tk.END)
                    phone_entry.insert(0, row[6])

                    # Load and set allergen checkboxes
                    allergens_str = row[7]
                    selected_allergens = allergens_str.split(",") if allergens_str else []

                    # Clear current checkbox selections
                    for var in environmental_allergen_vars.values():
                        var.set(False)
                    for var in venom_allergen_vars.values():
                        var.set(False)

                    #Set checkboxes based on loaded data:
                    if selected_allergens is not None:
                        for allergen in selected_allergens:
                            allergen = allergen.strip()
                            if allergen in environmental_allergen_vars:
                                environmental_allergen_vars[allergen].set(True)
                            elif allergen in venom_allergen_vars:
                                venom_allergen_vars[allergen].set(True)

                    # Set the Vial Type based on which allergens are selected
                    if any(allergen in venom_allergens for allergen in selected_allergens):
                        vial_type_var.set("Venom")
                    elif any(allergen in environmental_allergens for allergen in selected_allergens):
                        vial_type_var.set("Environmental")
                    else:
                        vial_type_var.set("Environmental")

                    load_window.destroy()  # Close the load window
                    return

            messagebox.showwarning("Patient Not Found", f"No patient found with Name and DOB: {selected_patient_str}")

        except FileNotFoundError:
            messagebox.showerror("Error", "patient_data.xlsx not found!")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")

    # --- Create the Load Patient Window ---
    load_window = tk.Toplevel(root)
    load_window.title("Load Patient")

    # Patient Selection Dropdown
    patient_select_label = ttk.Label(load_window, text="Select Patient (Name DOB):")
    patient_select_label.pack(padx=10, pady=5)

    patient_select_var = tk.StringVar()
    patient_select_combo = ttk.Combobox(load_window, textvariable=patient_select_var)
    patient_select_combo.pack(padx=10, pady=5)

    # Populate the dropdown with existing patient names and DOBs
    try:
        workbook = load_workbook(filename="patient_data.xlsx", read_only=True)
        sheet = workbook["Sheet1"]
        patient_strings = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                name = row[0]
                dob_str = row[1]
                # Format the DOB string for the dropdown
                dob_formatted = datetime.datetime.strptime(dob_str, "%m-%d-%Y").strftime("%m-%d-%Y")
                patient_strings.append(f"{name} {dob_formatted}")
            except ValueError:
                # Handle cases where the date in Excel might be invalid
                continue
        patient_select_combo['values'] = patient_strings

    except FileNotFoundError:
        messagebox.showinfo("No Patients", "No patient data found.  Add patients first.")
        load_window.destroy()
        return
    except Exception as e:
        messagebox.showerror("Error", f"Error loading patient list: {e}")
        load_window.destroy()
        return

    # Load Button
    load_button = ttk.Button(load_window, text="Load", command=load_selected_patient)
    load_button.pack(pady=10)


def generate_prescription():
    """Generates the prescription text."""
    mode = vial_type_var.get()
    patient_name = patient_name_entry.get()
    mrn = mrn_entry.get()
    address = address_entry.get()
    city = city_entry.get()
    state = state_entry.get()
    phone = phone_entry.get()

    if not patient_name or not mrn:
        messagebox.showerror("Error", "Please enter patient name and MRN.")
        return

    try:
        dob = dob_entry.get_date()
    except ValueError:
        messagebox.showerror("Error", "Invalid date of birth entered.")
        return

    patient_data = {
        'patient_name': patient_name,
        'dob': dob,
        'mrn': mrn,
        'address': address,
        'city': city,
        'state': state,
        'phone': phone
    }

    # --- Collect Selected Allergens ---
    if mode == "Environmental":
        selected_allergens = [
            allergen for allergen, var in environmental_allergen_vars.items() if var.get()
        ]
    elif mode == "Venom":
        selected_allergens = [
            allergen for allergen, var in venom_allergen_vars.items() if var.get()
        ]
    else:
        result_label.config(text="Invalid vial type selected.")
        return
    # Convert to comma-separated string
    selected_allergens_str = ", ".join(selected_allergens)

    save_patient_data(patient_data, selected_allergens_str)  # Pass the allergens

    prescription_text = f"Patient Name: {patient_name}\n"
    prescription_text += f"Date of Birth: {dob.strftime('%m-%d-%Y')}\n"
    prescription_text += f"MRN: {mrn}\n"
    prescription_text += f"Address: {address}\n"
    prescription_text += f"City: {city}, {state}\n"
    prescription_text += f"Phone: {phone}\n"
    prescription_text += f"Vial Type: {mode}\n"
    prescription_text += "Allergens:\n"

    if not selected_allergens:
        prescription_text += "  (No allergens selected)\n"
    else:
        for allergen in selected_allergens:
            prescription_text += f"  - {allergen}\n"

    result_label.config(text=prescription_text)


def update_allergen_options(*args):
    """Updates the visible allergen checkboxes based on vial type."""
    mode = vial_type_var.get()

    if mode == "Environmental":
        for widget in venom_allergen_frame.winfo_children():
            widget.grid_remove()
        for frame in environmental_allergen_frames:
            frame.grid()

    elif mode == "Venom":
        for frame in environmental_allergen_frames:
            frame.grid_remove()
        for i, allergen in enumerate(venom_allergens):
            venom_allergen_checkboxes[i].grid()

    else:
        for frame in environmental_allergen_frames:
            frame.grid_remove()
        for widget in venom_allergen_frame.winfo_children():
            widget.grid_remove()


def clear_fields():
    """Clears all input fields and checkbox selections."""
    patient_name_entry.delete(0, tk.END)
    dob_entry.set_date(datetime.date.today())
    mrn_entry.delete(0, tk.END)
    address_entry.delete(0, tk.END)
    city_entry.delete(0, tk.END)
    state_entry.delete(0, tk.END)
    phone_entry.delete(0, tk.END)
    # Uncheck all checkboxes
    for var in environmental_allergen_vars.values():
        var.set(False)
    for var in venom_allergen_vars.values():
        var.set(False)
    result_label.config(text="")  # Clear result label


# --- Main Application Window ---
root = tk.Tk()
root.title("Allergy Shot Vial Prescription Generator")
# --- Canvas and Scrollbar ---

main_frame = ttk.Frame(root)  # Create a main frame
main_frame.pack(fill=tk.BOTH, expand=1)  # Pack it to fill the window

my_canvas = tk.Canvas(main_frame)  # Canvas within the main_frame
my_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=1)

my_scrollbar = ttk.Scrollbar(main_frame, orient=tk.VERTICAL, command=my_canvas.yview)  # Create scrollbar
my_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)  # Pack it to the right side.

my_canvas.configure(yscrollcommand=my_scrollbar.set)  # Configure the canvas
my_canvas.bind('<Configure>', lambda e: my_canvas.configure(scrollregion=my_canvas.bbox("all")))  # bind configure

# --- Second Frame (inside Canvas) ---
second_frame = ttk.Frame(my_canvas)
my_canvas.create_window((0, 0), window=second_frame, anchor="nw")  # Add that new frame TO the canvas.

# --- Patient Information ---
patient_info_frame = ttk.LabelFrame(second_frame, text="Patient Information")  # Put it in second_frame
patient_info_frame.grid(row=0, column=0, padx=10, pady=10, sticky="ew")  # Grid it.

# Patient Name
patient_name_label = ttk.Label(patient_info_frame, text="Patient Name:")
patient_name_label.grid(row=0, column=0, padx=5, pady=5, sticky="w")
patient_name_entry = ttk.Entry(patient_info_frame)
patient_name_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

# Date of Birth
dob_label = ttk.Label(patient_info_frame, text="Date of Birth:")
dob_label.grid(row=1, column=0, padx=5, pady=5, sticky="w")
dob_entry = DateEntry(patient_info_frame, width=12, background='darkblue', foreground='white',
                      borderwidth=2, date_pattern='m-d-Y')
dob_entry.grid(row=1, column=1, padx=5, pady=5, sticky="ew")

# Medical Record Number
mrn_label = ttk.Label(patient_info_frame, text="MRN:")
mrn_label.grid(row=2, column=0, padx=5, pady=5, sticky="w")
mrn_entry = ttk.Entry(patient_info_frame)
mrn_entry.grid(row=2, column=1, padx=5, pady=5, sticky="ew")

# Street Address
address_label = ttk.Label(patient_info_frame, text="Street Address:")
address_label.grid(row=3, column=0, padx=5, pady=5, sticky="w")
address_entry = ttk.Entry(patient_info_frame)
address_entry.grid(row=3, column=1, padx=5, pady=5, sticky="ew")

# City
city_label = ttk.Label(patient_info_frame, text="City:")
city_label.grid(row=4, column=0, padx=5, pady=5, sticky="w")
city_entry = ttk.Entry(patient_info_frame)
city_entry.grid(row=4, column=1, padx=5, pady=5, sticky="ew")

# State
state_label = ttk.Label(patient_info_frame, text="State:")
state_label.grid(row=5, column=0, padx=5, pady=5, sticky="w")
state_entry = ttk.Entry(patient_info_frame)
state_entry.grid(row=5, column=1, padx=5, pady=5, sticky="ew")

# Phone Number
phone_label = ttk.Label(patient_info_frame, text="Phone Number:")
phone_label.grid(row=6, column=0, padx=5, pady=5, sticky="w")
phone_entry = ttk.Entry(patient_info_frame)
phone_entry.grid(row=6, column=1, padx=5, pady=5, sticky="ew")

# --- Vial Type Selection and Load Patient Button---
vial_type_label = ttk.Label(second_frame, text="Vial Type:")  # Put in second_frame
vial_type_label.grid(row=1, column=0, padx=5, pady=5, sticky="w")
vial_type_var = tk.StringVar(value="Environmental")
vial_type_var.trace_add("write", update_allergen_options)

vial_type_combo = ttk.Combobox(second_frame, textvariable=vial_type_var,  # Put in second_frame
                               values=["Environmental", "Venom"])
vial_type_combo.grid(row=1, column=0, padx=5, pady=5)  # Grid in row 1, column 0

load_patient_button = ttk.Button(second_frame, text="Load Patient", command=load_patient)  # Put in second_frame
load_patient_button.grid(row=1, column=0, padx=5, pady=5, sticky="e")  # Grid, sticky to east
# --- Environmental Allergen Checkboxes ---
environmental_allergen_frames = []
environmental_allergen_vars = {}

# --- Mold Group ---
mold_frame = ttk.LabelFrame(second_frame, text="Mold")  # second_frame
mold_frame.grid(row=2, column=0, padx=5, pady=5, sticky="ew")  # Adjusted row
environmental_allergen_frames.append(mold_frame)
mold_allergens = ["Aspergillus", "Alternaria", "Cladosporium", "Penicillium"]
for i, allergen in enumerate(mold_allergens):
    var = tk.BooleanVar()
    environmental_allergen_vars[allergen] = var
    checkbox = ttk.Checkbutton(mold_frame, text=allergen, variable=var)
    checkbox.grid(row=i // 2, column=i % 2, padx=5, pady=2, sticky="w")  # No changes here.

# --- Tree Group ---
tree_frame = ttk.LabelFrame(second_frame, text="Tree")  # second_frame
tree_frame.grid(row=3, column=0, padx=5, pady=5, sticky="ew")  # Adjusted row
environmental_allergen_frames.append(tree_frame)
tree_allergens = ["Ash", "Birch (Oak)", "Cedar", "Hackberry (Elm)", "Maple", "Sycamore", "Walnut (Pecan)",
                  "Willow (Cottonwood)", "Mulberry"]
for i, allergen in enumerate(tree_allergens):
    var = tk.BooleanVar()
    environmental_allergen_vars[allergen] = var
    checkbox = ttk.Checkbutton(tree_frame, text=allergen, variable=var)
    checkbox.grid(row=i // 2, column=i % 2, padx=5, pady=2, sticky="w")

# --- Grass Group ---
grass_frame = ttk.LabelFrame(second_frame, text="Grass")  # second_frame
grass_frame.grid(row=4, column=0, padx=5, pady=5, sticky="ew")  # Adjusted row
environmental_allergen_frames.append(grass_frame)
grass_allergens = ["Timothy", "Johnson", "Bermuda"]
for i, allergen in enumerate(grass_allergens):
    var = tk.BooleanVar()
    environmental_allergen_vars[allergen] = var
    checkbox = ttk.Checkbutton(grass_frame, text=allergen, variable=var)
    checkbox.grid(row=i // 2, column=i % 2, padx=5, pady=2, sticky="w")

# --- Weed Group ---
weed_frame = ttk.LabelFrame(second_frame, text="Weed")  # second_frame
weed_frame.grid(row=5, column=0, padx=5, pady=5, sticky="ew")  # Adjusted row
environmental_allergen_frames.append(weed_frame)
weed_allergens = ["Cocklebur", "Yellow Dock (Sheep Sorrel)", "Kochia (Firebush)", "Lamb's Quarter", "Mugwort",
                  "Pigweed", "English Plantain", "Russian Thistle", "Ragweed"]
for i, allergen in enumerate(weed_allergens):
    var = tk.BooleanVar()
    environmental_allergen_vars[allergen] = var
    checkbox = ttk.Checkbutton(weed_frame, text=allergen, variable=var)
    checkbox.grid(row=i // 2, column=i % 2, padx=5, pady=2, sticky="w")

# --- Other Group ---
other_frame = ttk.LabelFrame(second_frame, text="Other")  # second_frame
other_frame.grid(row=6, column=0, padx=5, pady=5, sticky="ew")  # Adjusted row
environmental_allergen_frames.append(other_frame)
other_allergens = ["Cat", "Dog - UF", "Dog - Epithelium", "Mouse", "Horse", "Amer. Cockroach", "Ger. Cockroach",
                   "Dust Mite Mix"]
for i, allergen in enumerate(other_allergens):
    var = tk.BooleanVar()
    environmental_allergen_vars[allergen] = var
    checkbox = ttk.Checkbutton(other_frame, text=allergen, variable=var)
    checkbox.grid(row=i // 2, column=i % 2, padx=5, pady=2, sticky="w")

# --- Venom Allergen Checkboxes ---
venom_allergen_frame = ttk.LabelFrame(second_frame, text="Venom Allergens")  # second_frame
venom_allergen_frame.grid(row=7, column=0, columnspan=2, padx=5, pady=5, sticky="ew")  # Adjusted row

venom_allergens = ["Honey Bee", "Yellow Jacket", "Yellow Faced Hornet", "White Faced Hornet", "Wasp"]
venom_allergen_vars = {}
venom_allergen_checkboxes = []

for i, allergen in enumerate(venom_allergens):
    var = tk.BooleanVar()
    venom_allergen_vars[allergen] = var
    checkbox = ttk.Checkbutton(venom_allergen_frame, text=allergen, variable=var)
    checkbox.grid(row=i // 2, column=i % 2, padx=5, pady=2, sticky="w")
    venom_allergen_checkboxes.append(checkbox)

# Initially hide venom allergens
for widget in venom_allergen_frame.winfo_children():
    widget.grid_remove()

# --- Generate, Clear and Load Buttons ---
generate_button = ttk.Button(second_frame, text="Generate Prescription", command=generate_prescription)  # second_frame
generate_button.grid(row=8, column=0, padx=5, pady=10, sticky="w")

clear_button = ttk.Button(second_frame, text="Clear Fields", command=clear_fields)  # second_frame
clear_button.grid(row=8, column=0, padx=5, pady=10, sticky="e")  # Added clear button

# --- Result Label ---
result_label = ttk.Label(second_frame, text="", wraplength=600)  # Increased wraplength
result_label.grid(row=9, column=0, padx=5, pady=5)  # Adjusted row

root.mainloop()
